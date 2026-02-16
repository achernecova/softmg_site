import json
import logging
import re
from urllib.parse import urljoin

import allure
import requests
from allure_commons.types import AttachmentType
from faker import Faker
from openpyxl import Workbook

from config import BASE_URL, PageConfig

logger = logging.getLogger(__name__)


class ApiHelper:
    def __init__(self):
        self.name_data = Faker()
        self.session = requests.Session()
        self.correct_email = self.name_data.email()

    def search(self):
        with allure.step("Отправляем запрос поиска"):
            response = self.session.get(
                url=f"{BASE_URL}/api/v2/search",
                params={"q": "тестирование", "limit": 5},
                allow_redirects=False,
            )
            logger.info("Отправили запрос на поиск")

            self.log_request_and_response(response.request, response)
            assert (
                    "PHPSESSID" in response.cookies
            ), "Сервер не вернул куки PHPSESSID"
            print(response.cookies.get_dict())
            return response.cookies.get_dict()

    def articles_search(self):
        with allure.step("Отправляем запрос поиска"):
            response = self.session.get(
                url=f"{BASE_URL}/api/v2/articles",
                params={"tag": "разработка", "limit": 5},
                allow_redirects=False,
            )
            logger.info("Отправили запрос на поиск")

            self.log_request_and_response(response.request, response)
            assert (
                    "PHPSESSID" in response.cookies
            ), "Сервер не вернул куки PHPSESSID"
            print(response.cookies.get_dict())
            return response.cookies.get_dict()

    def search_data_examples_with_tag_meditsina(self):
        with allure.step("Отправляем запрос с тегом meditsina"):
            response = self.session.get(
                url=f"{BASE_URL}/api/v2/cases",
                params={"tag": "meditsina", "limit": 5},
                allow_redirects=False,
            )
            logger.info("Отправили запрос на поиск")

            self.log_request_and_response(response.request, response)

            # Преобразуем строку JSON в словарь Python
            body = self.transformation_json(response)

            # Доступ к полям сразу
        with allure.step("Забираем кол-во записей, мета-данные и имя"):
            response_title = body['data']
            response_meta = body['meta']['count']
            names = [project.get('name', '') for project in response_title]

            return response.status_code, len(response_title), response.json(), response_meta, names

    @staticmethod
    def log_request_and_response(request, response):
        request_info = f"Request Method: {request.method}\nRequest URL: {request.url}\n"

        if request.body:
            request_info += f"Request Body: {request.body}\n"
        if request.headers:
            request_info += f"Request Headers: {json.dumps(dict(request.headers), indent=4, sort_keys=True)}\n"
        allure.attach(
            body=request_info,
            name="Request Info",
            attachment_type=AttachmentType.TEXT,
            extension="txt",
        )

        allure.attach(
            body=str(response.status_code),
            name="Response Status Code",
            attachment_type=AttachmentType.TEXT,
            extension="txt",
        )

        resp_body_json = response.json()
        resp_body_str = json.dumps(resp_body_json, indent=4, sort_keys=True)

        allure.attach(
            body=resp_body_str,
            name="Response Body",
            attachment_type=AttachmentType.JSON,
            extension="json",
        )

        allure.attach(
            body=str(response.cookies),
            name="Response Cookies",
            attachment_type=AttachmentType.JSON,
            extension="json",
        )

        logger.info(resp_body_str)
        logger.info(request_info)

    def add_request_with_only_input_all_fields(self, value):
        """
        Отправка заявки с корректно заполненными данными
        :param value: названия форм с которых будет отправлена заявка
        :return:  возвращаем статус-код
        """
        with allure.step(f"Отправляем заявку с формы {value}"):
            response = requests.post(
                url=BASE_URL + "api/v2/feedback/add",
                data={"feedback_name": value,
                      "email": self.correct_email,
                      "name": self.name_data.name_male(),
                      "phone": self.name_data.random_int(10000000000, 99999999999),
                      "description": self.name_data.text(500)},
                headers={"Accept": "application/json"},
                allow_redirects=False,
            )
            logger.info(f"Отправили заявку из формы {value}")
            self.log_request_and_response(response.request, response)
        return response.status_code

    def add_request_with_not_correct_email(self, value):
        """
        Отправка заявки с некорректным email
        :param value: названия форм с которых будет отправлена заявка
        :return: возвращаем статус-код, ошибку и указываемую почту
        """
        email_data = self.name_data.text(max_nb_chars=20)
        with allure.step(f"Отправляем заявку с формы {value}"):
            response = requests.post(
                url=BASE_URL + "api/v2/feedback/add",
                data={"feedback_name": {value},
                      "email": email_data
                      },
                headers={"Accept": "application/json"},
                allow_redirects=False,
            )
            logger.info(f"Отправили заявку из формы {value}")
            self.log_request_and_response(response.request, response)

            # Преобразование JSON в словарь
            body = self.transformation_json(response)

            # Доступ к полям сразу
            error_title = body['errors'][0]['title']
        return response.status_code, error_title, email_data

    def add_request_in_request_with_one_char_in_email(self, value):
        """
        Отправка заявки с некорректным email - кол-во символов в имени меньше 3‑х
        :param value: названия форм с которых будет отправлена заявка
        :return: возвращаем статус-код
        """
        email_data = "1@test.com"
        with allure.step(f"Отправляем заявку с формы {value}"):
            response = requests.post(
                url=BASE_URL + "api/v2/feedback/add",
                data={"feedback_name": {value},
                      "email": email_data
                      },
                headers={"Accept": "application/json"},
                allow_redirects=False,
            )
            logger.info(f"Отправили заявку из формы {value}")
            self.log_request_and_response(response.request, response)
        return response.status_code

    def add_request_with_all_fields_empty(self, value):
        """
        Отправка заявки с незаполненными полями
        :param value: названия форм с которых будет отправлена заявка
        :return: возвращаем статус-код и ошибку
        """
        with allure.step(f"Отправляем заявку с формы {value}"):
            response = requests.post(
                url=BASE_URL + "api/v2/feedback/add",
                data={"feedback_name": {value}
                      },
                headers={"Accept": "application/json"},
                allow_redirects=False,
            )
            logger.info(f"Отправили заявку из формы {value}")
            self.log_request_and_response(response.request, response)

            # Преобразование JSON в словарь
            body = self.transformation_json(response)

            # Доступ к полям сразу
            error_title = body['errors'][0]['title']
        return response.status_code, error_title

    def add_request_in_form_with_not_correct_phone(self, value):
        """
        Отправка заявки с кривым номером телефона
        :param value: названия форм с которых будет отправлена заявка
        :return: возвращаем статус-код, текст ошибки и ожидаемый формат
        """
        phone_number = self.name_data.random_int(100000, 999999)
        with allure.step(f"Отправляем заявку с формы {value}"):
            response = requests.post(
                url=BASE_URL + "api/v2/feedback/add",
                data={"feedback_name": value,
                      "email": self.correct_email,
                      "phone": phone_number,
                      },
                headers={"Accept": "application/json"},
                allow_redirects=False,
            )
            logger.info(f"Отправили заявку из формы {value}")
            self.log_request_and_response(response.request, response)

            # Преобразование JSON в словарь
            body = self.transformation_json(response)

            # Доступ к полю title
            error_title = body['errors'][0]['title']

            # Через регулярки сравниваем полученную ошибку из поля title с заданным значением
            pattern = r'^The phone number ""\w+"" is not a valid phone number\.$'
            error_text = re.match(pattern, error_title)

        return response.status_code, error_text, error_title

    @staticmethod
    def transformation_json(value):
        try:
            body = json.loads(value.content)
        except json.JSONDecodeError as e:
            logger.error(f"Ошибка парсинга JSON: {e}")
            return False
        return body

    def add_request_in_form_with_tg_email_wa_phone(self, value):
        """
        Отправка заявки с корректно заполненными данными
        :param value: названия форм с которых будет отправлена заявка
        :return: возвращаем статус-код
        """

        telegram_data = "@" + self.name_data.name()
        phone_number = self.name_data.random_int(10000000000, 99999999999)
        with allure.step(f"Отправляем заявку с формы {value}"):
            response = requests.post(
                url=BASE_URL + "api/v2/feedback/add",
                data={"feedback_name": value,
                      "email": self.correct_email,
                      "name": self.name_data.name_male(),
                      "phone": phone_number,
                      "telegram": telegram_data,
                      "whatsapp": phone_number,
                      "description": self.name_data.text(500)},
                headers={"Accept": "application/json"},
                allow_redirects=False,
            )
            logger.info(f"Отправили заявку из формы {value}")
            self.log_request_and_response(response.request, response)
        return response.status_code

    def add_request_in_form_with_with_exceeding_characters_in_descr(self, value):
        """
        Отправка заявки с превышением символов в поле description
        :param value: названия форм с которых будет отправлена заявка
        :return: возвращаем статус-код
        """
        with allure.step(f"Отправляем заявку с формы {value}"):
            response = requests.post(
                url=BASE_URL + "api/v2/feedback/add",
                data={"feedback_name": value,
                      "email": self.correct_email,
                      "description": self.name_data.text(1000)},
                headers={"Accept": "application/json"},
                allow_redirects=False,
            )
            logger.info(f"Отправили заявку из формы {value}")
            self.log_request_and_response(response.request, response)
        return response.status_code

    def get_links_in_main_page(self, value):
        with allure.step(f"Отправляем запрос на получение данных с {value} страницы"):
            response = self.session.get(
                url=f"{BASE_URL}/api/v2/{value}",
                allow_redirects=False,
            )
            logger.info("Отправили запрос на поиск")

            self.log_request_and_response(response.request, response)
            return response

    def extract_links(self, data, links_set):
        """
        Метод для рекурсивного сбора ссылок с учётом типа данных
        :param data: получаемый респонс
        :param links_set: список урлов
        """
        if isinstance(data, dict):
            # Если это категория "cases", добавляем путь /examples/
            if "cases" in data.get("props", {}):
                for case in data["props"]["cases"]:
                    link = case.get("link", "").strip("/")
                    if link:
                        full_path = "/examples/" + link
                        if not any(full_path in existing_link for existing_link in links_set):
                            links_set.add(full_path)

            # Если это категория "articles", используем slug для формирования ссылки
            elif "articles" in data.get("props", {}):
                for article in data["props"]["articles"]:
                    slug = article.get("slug", "").strip("/")
                    if slug:
                        full_path = "/article/" + slug
                        if not any(full_path in existing_link for existing_link in links_set):
                            links_set.add(full_path)

            # Прямые ссылки (для прочих случаев)
            elif "link" in data:
                link = data["link"].strip("/")
                if link and not any(link in existing_link for existing_link in links_set):
                    links_set.add(link)

            # Продолжаем обход остальных свойств с исключением блока tags
            for key, value in data.items():
                if key != "tags":
                    self.extract_links(value, links_set)

        elif isinstance(data, list):
            # Если это список, проходим по каждому элементу
            for item in data:
                self.extract_links(item, links_set)

    def fetch_and_test_links(self, url):
        with allure.step(f"Отправляем запрос на получение данных с {url} страницы"):
            response = self.get_links_in_main_page(url)

        with allure.step("Проверяем успешность запроса"):
            if response.status_code != 200:
                raise Exception(f"Ошибка при получении данных: {response.status_code}")

        with allure.step("Парсим JSON-ответ и собираем все ссылки через рекурсию"):
            data = response.json()
            unique_links = set()
            self.extract_links(data, unique_links)

        with allure.step("Формируем полные URL для каждой уникальной ссылки"):
            full_urls = [urljoin(BASE_URL, link) for link in unique_links]

        with allure.step("Проверяем каждую ссылку отдельным GET-запросом"):
            results = {}
            for url in full_urls:
                sub_response = self.session.get(url, allow_redirects=True)
                results[url] = sub_response.status_code

            return results

    @staticmethod
    def get_api_urls():
        """
        Возвращает список кортежей (api_url, url), которые нужны для дальнейших запросов.
        Пропускает страницы, у которых отсутствует или пустой api_url.
        """
        pages_config = PageConfig().pages
        return [(page["api_url"], page["url_page"]) for page in pages_config.values() if
                "api_url" in page and page["api_url"] != ""]

    @staticmethod
    def extract_section_types(response_data):
        """
        Извлекает типы блоков ('type') из раздела 'sections'.
        """
        try:
            return [section['type'] for section in response_data['sections']]
        except KeyError:
            return []

    def fetch_and_process_pages(self):
        """
        Выполняет запросы ко всем API URL и собирает типы блоков.
        Возвращает словарное представление с результатами и список ошибок.
        """
        api_urls = self.get_api_urls()
        results = {}
        errors = []

        for api_url, url in api_urls:
            try:
                response = requests.get(api_url)
                if response.status_code == 200:
                    data = response.json()
                    types = self.extract_section_types(data)
                    results[url] = (api_url, types)
                elif response.status_code == 404:
                    errors.append((url, "Страница не найдена (404)"))
                else:
                    errors.append((url, f"Ошибка при обработке URL {url}, статус-код: {response.status_code}"))
            except Exception as e:
                errors.append((url, f"Ошибка при обработке URL {url}: {e}"))

        return results, errors

    @staticmethod
    def save_to_excel(results):
        """
        Преобразуем словарь с результатами в Excel и сохраняет.
        """
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "URL and API URL"

        # Записываем данные на первую страницу
        for idx, (url, (api_url, types)) in enumerate(results.items(), start=1):
            # Добавляем гиперссылку для url
            sheet1.cell(row=idx, column=1).value = url
            sheet1.cell(row=idx, column=1).hyperlink = url
            sheet1.cell(row=idx, column=1).style = "Hyperlink"  # Применяем стиль гиперссылки

            # Добавляем гиперссылку api_url
            sheet1.cell(row=idx, column=2).value = api_url
            sheet1.cell(row=idx, column=2).hyperlink = api_url
            sheet1.cell(row=idx, column=2).style = "Hyperlink"  # Применяем стиль гиперссылки

            for col_idx, block_type in enumerate(types, start=3):
                sheet1.cell(row=idx, column=col_idx, value=block_type)

        # Создаем вторую страницу
        sheet2 = workbook.create_sheet(title="Blocks and URLs")

        # Собираем данные о блоках и URL
        block_urls = {}
        for url, (api_url, types) in results.items():
            for block_type in types:
                if block_type not in block_urls:
                    block_urls[block_type] = []
                block_urls[block_type].append(url)

        # Записываем данные на вторую страницу
        for idx, (block_type, urls) in enumerate(block_urls.items(), start=1):
            sheet2.cell(row=idx, column=1).value = block_type
            for col_idx, url in enumerate(urls, start=2):
                # sheet2.cell(row=idx, column=col_idx, value=url)
                sheet2.cell(row=idx, column=col_idx).value = url
                sheet2.cell(row=idx, column=col_idx).hyperlink = url
                sheet2.cell(row=idx, column=col_idx).style = "Hyperlink"

        # Сохраняем файл
        excel_file = 'block_types.xlsx'
        try:
            workbook.save(excel_file)
            print(f"Данные успешно сохранены в файл {excel_file}.")
        except Exception as e:
            print(f"Ошибка при сохранении файла: {e}")
